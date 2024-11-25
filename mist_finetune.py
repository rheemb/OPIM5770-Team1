import fitz  # PyMuPDF
import os
import time
import json
import torch
import numpy as np
import logging
from sentence_transformers import SentenceTransformer
import faiss
from datasets import Dataset
from transformers import AutoModelForCausalLM, AutoTokenizer, TrainingArguments
from peft import LoraConfig, get_peft_model
from trl import SFTTrainer
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import datetime

# Initialize logging
log_filename = "chatbot_finetune.log"
logging.basicConfig(filename=log_filename, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger()
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

# Record the start time
start_time = time.time()

# Step 1: PDF Preprocessing
def extract_text_from_pdf(pdf_path):
    try:
        doc = fitz.open(pdf_path)
        text = ""
        for page in doc:
            text += page.get_text()
        return text
    except Exception as e:
        logger.error(f"Error reading {pdf_path}: {str(e)}")
        return ""

def preprocess_pdfs(folder_path):
    all_text = ""
    for filename in os.listdir(folder_path):
        if filename.lower().endswith(".pdf"):
            pdf_path = os.path.join(folder_path, filename)
            logger.info(f"Extracting text from: {pdf_path}")
            text = extract_text_from_pdf(pdf_path)
            all_text += text + "\n"
    return all_text

# Step 2: Text Splitting for Dataset Creation
def create_training_samples(text, max_length=512):
    words = text.split()
    samples = []
    for i in range(0, len(words), max_length // 2):
        input_text = " ".join(words[i:i + max_length // 2])
        output_text = " ".join(words[i + max_length // 2:i + max_length])
        samples.append({"input": input_text, "output": output_text})
    return samples

# Step 3: Prepare Dataset for Fine-Tuning
def prepare_dataset(samples):
    return [{"prompt": sample["input"], "response": sample["output"]} for sample in samples]

# Step 4: Fine-Tuning Process
def fine_tune_model(model, tokenizer, train_dataset, eval_dataset):
    try:
        peft_config = LoraConfig(
            lora_alpha=16,
            lora_dropout=0.1,
            r=64,
            bias="none",
            task_type="CAUSAL_LM"
        )
        model = get_peft_model(model, peft_config)

        training_args = TrainingArguments(
            output_dir="./fine_tuned_mistral_mist",
            evaluation_strategy="epoch",
            learning_rate=2e-5,
            per_device_train_batch_size=2,
            num_train_epochs=10, #10,30
            weight_decay=0.01,
            save_total_limit=2,
            save_strategy="epoch",
            logging_steps=10, #50 ( if you want to check detail training step losses)
            fp16=torch.cuda.is_available(),
            gradient_accumulation_steps=4,
            load_best_model_at_end=True,
            metric_for_best_model="loss",
            greater_is_better=False,
        )

        trainer = SFTTrainer(
            model=model,
            peft_config=peft_config,
            max_seq_length=512,
            tokenizer=tokenizer,
            args=training_args,
            train_dataset=train_dataset,
            eval_dataset=eval_dataset,
            dataset_text_field="prompt",
        )

        logger.info("Starting fine-tuning...")
        trainer.train()
        logger.info("Fine-tuning completed successfully")
        return model
    except Exception as e:
        logger.error(f"Error during fine-tuning: {str(e)}")
        raise

# Step 5: Generate Responses (without retrieval)
def generate_response(prompt, model, tokenizer, max_length=250):
    try:
        inputs = tokenizer(prompt, return_tensors="pt").to(model.device)
        outputs = model.generate(
            **inputs,
            max_length=max_length,
            temperature=0.7,
            top_p=0.9,
            do_sample=True,
            num_return_sequences=1,
            eos_token_id=tokenizer.eos_token_id,
            pad_token_id=tokenizer.eos_token_id,
            no_repeat_ngram_size=3,
            early_stopping=True
        )
        return tokenizer.decode(outputs[0], skip_special_tokens=True)
    except Exception as e:
        logger.error(f"Error generating response: {str(e)}")
        return "I'm sorry, but I couldn't generate a response at this time."

# Step 6: Excel Logging Function
def save_qa_to_excel(qa_pairs, filename="qa_log.xlsx"):
    try:
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
            start_row = ws.max_row + 1
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Q&A Log"
            headers = ["Timestamp", "Question", "Answer"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            start_row = 2

        for row, (timestamp, question, answer) in enumerate(qa_pairs, start=start_row):
            ws.cell(row=row, column=1, value=timestamp)
            ws.cell(row=row, column=2, value=question)
            ws.cell(row=row, column=3, value=answer)

        wb.save(filename)
        logger.info(f"Q&A pairs saved to {filename}")
    except Exception as e:
        logger.error(f"Error saving Q&A pairs to Excel: {str(e)}")

# Main Execution
if __name__ == "__main__":
    try:
        # Specify the path to your PDFs
        pdf_folder_path = "/root/Capstone Chatbots/mis_mod/data/Docs"
        fine_tuned_model_path = "./fine_tuned_mistral_mist"
        base_model_path = "/root/Capstone Chatbots/mis_mod/mistral/7B-Instruct-v0.3"

        # Extract text from PDFs
        logger.info("Preprocessing PDFs...")
        pdf_text = preprocess_pdfs(pdf_folder_path)

        # Create training samples
        logger.info("Creating training samples...")
        samples = create_training_samples(pdf_text)

        # Prepare dataset
        logger.info("Preparing dataset...")
        dataset = prepare_dataset(samples)

        # Create Hugging Face Dataset
        logger.info("Creating Hugging Face dataset...")
        hf_dataset = Dataset.from_dict({
            "prompt": [d["prompt"] for d in dataset],
            "response": [d["response"] for d in dataset]
        })

        # Split into train and evaluation sets
        logger.info("Splitting dataset into train and evaluation sets...")
        train_test_split = hf_dataset.train_test_split(test_size=0.2, shuffle=True, seed=42)
        train_dataset = train_test_split['train']
        eval_dataset = train_test_split['test']
        logger.info(f"Training dataset size: {len(train_dataset)}")
        logger.info(f"Evaluation dataset size: {len(eval_dataset)}")

        # Load or fine-tune the model
        if os.path.exists(fine_tuned_model_path):
            logger.info("Loading pre-fine-tuned model...")
            tokenizer = AutoTokenizer.from_pretrained(fine_tuned_model_path)
            model = AutoModelForCausalLM.from_pretrained(
                fine_tuned_model_path,
                torch_dtype=torch.float16 if torch.cuda.is_available() else torch.float32,
                device_map="auto"
            )
        else:
            logger.info("Fine-tuned model not found. Loading base model and starting fine-tuning...")
            tokenizer = AutoTokenizer.from_pretrained(base_model_path)
            if tokenizer.pad_token is None:
                tokenizer.pad_token = tokenizer.eos_token
            model = AutoModelForCausalLM.from_pretrained(
                base_model_path,
                torch_dtype=torch.float16 if torch.cuda.is_available() else torch.float32,
                device_map="auto"
            )
            model.config.pad_token_id = tokenizer.pad_token_id

            # Fine-tune the model
            logger.info("Starting fine-tuning...")
            model = fine_tune_model(model, tokenizer, train_dataset, eval_dataset)
            
            # Save the fine-tuned model
            logger.info("Saving fine-tuned model...")
            model.save_pretrained(fine_tuned_model_path)
            tokenizer.save_pretrained(fine_tuned_model_path)

        logger.info("=== Model loaded. Chatbot is ready! ===")

        qa_pairs = []
        # Interactive loop
        while True:
            user_query = input("\nEnter your question (or type 'exit' to quit): ")
            if user_query.lower() == "exit":
                break
            
            response = generate_response(user_query, model, tokenizer)
            print(f"\nResponse: {response}")

            # Store Q&A pair
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            qa_pairs.append((timestamp, user_query, response))

        # Save Q&A pairs to Excel
        if qa_pairs:
            save_qa_to_excel(qa_pairs)
        else:
            logger.info("No Q&A pairs to save.")

    except Exception as e:
        logger.error(f"An error occurred: {str(e)}", exc_info=True)

    finally:
        logger.info("Thank you for using the chatbot.")
        end_time = time.time()
        total_time = end_time - start_time
        logger.info(f"Total execution time: {total_time:.2f} seconds")