import torch
from transformers import AutoModelForCausalLM, AutoTokenizer
import logging
import time
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import os
import re

# Initialize logging
log_filename = "chata_llama.log"
logging.basicConfig(filename=log_filename, level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger()

start_time = time.time()

def generate_response(query, model, tokenizer, max_new_tokens=500, temperature=0.3, top_p=0.95):
    logger.info(f"Received query: {query}")
    
    prompt = f"Human: {query}\n\nAssistant: Please provide a detailed explanation."
    
    inputs = tokenizer(prompt, return_tensors="pt").to(model.device)
    outputs = model.generate(
        **inputs,
        max_new_tokens=max_new_tokens,
        temperature=temperature,
        top_p=top_p,
        do_sample=True,
        eos_token_id=tokenizer.eos_token_id,
        pad_token_id=tokenizer.eos_token_id
    )
    
    response = tokenizer.decode(outputs[0], skip_special_tokens=True)
    answer = response.split("Assistant:")[-1].strip()
    
    logger.info(f"Generated answer: {answer[:100]}...")  
    return answer

def sanitize_for_excel(text):
    """Remove or replace characters that Excel doesn't support."""
    # Remove control characters
    text = ''.join(char for char in text if ord(char) >= 32)
    # Replace other problematic characters
    text = re.sub(r'[\000-\010]|[\013-\014]|[\016-\031]', '', text)
    # Limit the length of the text (Excel has a 32,767 character limit per cell)
    return text[:32000]  # Leave some margin for safety

def process_questions_from_excel(file_path, model, tokenizer):
    try:
        # Load the Excel file
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            logger.error(f"File {file_path} does not exist.")
            return

        # Iterate through the rows and process questions
        for row in range(2, ws.max_row + 1):  # Assuming questions start from row 2
            question_cell = ws.cell(row=row, column=1)
            if question_cell.value is None:
                continue  # Skip empty cells

            question = question_cell.value
            answer = generate_response(question, model, tokenizer)

            # Write the answer to the next cell
            answer_cell = ws.cell(row=row, column=2)
            answer_cell.value = sanitize_for_excel(answer)

        # Save the workbook
        wb.save(file_path)
        logger.info(f"Questions processed and answers saved to {file_path}")

    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    # Load the LLaMA model and tokenizer from local storage
    model_path = "/root/Capstone Chatbots/.venv/llama_mod/"
    model = AutoModelForCausalLM.from_pretrained(model_path)
    tokenizer = AutoTokenizer.from_pretrained(model_path)

    # Ensure the model is moved to the correct device (GPU if available)
    device = torch.device("cuda")
    model.to(device)
    model.config.pad_token_id = model.config.eos_token_id

    logger.info("\n=== Chatbot is ready! ===")

    # Process questions from Excel file
    excel_file_path = "/root/Capstone Chatbots/.venv/qstns.xlsx"  # Replace with your Excel file path
    process_questions_from_excel(excel_file_path, model, tokenizer)

    # Calculate and log total execution time
    end_time = time.time()
    total_time = end_time - start_time
    logger.info(f"Total execution time: {total_time:.2f} seconds")