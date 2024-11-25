import fitz  # PyMuPDF
import os
import time
import logging
import re
from sentence_transformers import SentenceTransformer
import faiss
import torch
from transformers import AutoModelForCausalLM, AutoTokenizer
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import datetime

# Initialize logging
log_filename = "chatbot.log"
logging.basicConfig(filename=log_filename, level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger()

# To also print logs to console
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

# Record the start time
start_time = time.time()

# ==============================
# Step 1: PDF Preprocessing and Text Extraction
# ==============================

def extract_text_from_pdf(pdf_path):
    """Extract text from a single PDF file."""
    try:
        doc = fitz.open(pdf_path)
        text = ""
        for page in doc:
            text += page.get_text()
        return text
    except Exception as e:
        logger.error(f"Error reading {pdf_path}: {e}")
        return ""

def preprocess_pdfs(folder_path):
    """Extract text from all PDFs in a specified folder."""
    all_text = ""
    for filename in os.listdir(folder_path):
        if filename.lower().endswith(".pdf"):
            pdf_path = os.path.join(folder_path, filename)
            logger.info(f"Extracting text from: {pdf_path}")
            text = extract_text_from_pdf(pdf_path)
            all_text += text + "\n"  # Add a newline to separate documents
    return all_text

# ==============================
# Step 2: Text Splitting and Chunking
# ==============================

def split_text_into_chunks(text, chunk_size=2500, overlap=150):
    """Split text into chunks with specified size and overlap."""
    chunks = []
    start = 0
    text_length = len(text)
    while start < text_length:
        end = start + chunk_size
        chunk = text[start:end]
        chunks.append(chunk)
        start += chunk_size - overlap
    return chunks

# ==============================
# Step 3: Creating a Retrieval System with FAISS
# ==============================

def create_faiss_index(chunks, embedding_model_name="sentence-transformers/all-MiniLM-L6-v2"):
    """Create a FAISS index from text chunks."""
    logger.info("Creating embeddings for chunks...")
    embedding_model = SentenceTransformer(embedding_model_name)
    embeddings = embedding_model.encode(chunks, convert_to_numpy=True, show_progress_bar=True)
    logger.info("Initializing FAISS index...")
    index = faiss.IndexFlatL2(embeddings.shape[1])
    index.add(embeddings)
    return index, embedding_model

def save_faiss_index(index, path="faiss_index.bin"):
    """Save the FAISS index to disk."""
    faiss.write_index(index, path)
    logger.info(f"FAISS index saved to {path}")

# ==============================
# Step 4: RAG Implementation
# ==============================
def generate_context_aware_response(query, faiss_index, pdf_chunks, embedding_model, model, tokenizer, max_new_tokens=1000, temperature=0.7, top_p=0.9):
    logger.info(f"Received query: {query}")
    
    # Embed the query
    query_embedding = embedding_model.encode([query], convert_to_numpy=True)
    
    # Search FAISS index for top-k relevant chunks
    k = 5  # Number of chunks to retrieve
    distances, indices = faiss_index.search(query_embedding, k)
    
    # Retrieve the relevant chunks
    relevant_chunks = [pdf_chunks[idx] for idx in indices[0]]
    context = "\n".join(relevant_chunks)
    
    # Create the prompt with context
    prompt = f"Based on the following context, please answer the question:\n\nContext: {context}\n\nQuestion: {query}\nAnswer:"
    
    logger.info(f"Generated prompt: {prompt[:100]}...")  # Log first 100 chars of prompt
    
    # Tokenize and generate response
    inputs = tokenizer(prompt, return_tensors="pt").to(model.device)
    outputs = model.generate(
        **inputs,
        # max_length=max_length,
        max_new_tokens=max_new_tokens,
        temperature=temperature,
        top_p=top_p,
        do_sample=True,
        eos_token_id=tokenizer.eos_token_id,
        pad_token_id=tokenizer.eos_token_id
    )
    
    response = tokenizer.decode(outputs[0], skip_special_tokens=True)
    answer = response.split("Answer:")[-1].strip()  # Extract the answer part
    logger.info(f"Generated answer: {answer[:100]}...")  # Log first 100 chars of answer
    
    return context, answer

# ==============================
# sanitizing the text of excel
# # ==============================
def sanitize_for_excel(text):
    """Remove or replace characters that Excel doesn't support."""
    # Remove control characters
    text = ''.join(char for char in text if ord(char) >= 32)
    # Replace other problematic characters
    text = re.sub(r'[\000-\010]|[\013-\014]|[\016-\031]', '', text)
    # Limit the length of the text (Excel has a 32,767 character limit per cell)
    return text[:32000]  # Leave some margin for safety

# ==============================
# Excel Logging Function
# ==============================
def save_qa_to_excel(qa_pairs, filename="qa_log_4.xlsx"):
    if os.path.exists(filename):
        # File exists, load it and append data
        wb = load_workbook(filename)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        # File doesn't exist, create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Q&A Log"
        # Add headers
        headers = ["Timestamp", "Question", "Context", "Answer"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        start_row = 2

    # Add Q&A pairs
    for row, (timestamp, question, context, answer) in enumerate(qa_pairs, start=start_row):
        ws.cell(row=row, column=1, value=timestamp)
        ws.cell(row=row, column=2, value=sanitize_for_excel(question))
        ws.cell(row=row, column=3, value=sanitize_for_excel(context))
        ws.cell(row=row, column=4, value=sanitize_for_excel(answer))
    
    # Save without adjusting column widths
    wb.save(filename)
    print(f"Q&A pairs saved to {filename}")

# ==============================
# Main Execution
# ==============================

if __name__ == "__main__":
    # Specify the path to your PDFs
    pdf_folder_path = "/root/Capstone Chatbots/mis_mod/data/Docs"

    # Extract text from PDFs
    logger.info("Preprocessing PDFs...")
    pdf_text = preprocess_pdfs(pdf_folder_path)

    # Split text into chunks
    logger.info("Splitting text into chunks...")
    pdf_chunks = split_text_into_chunks(pdf_text)
    logger.info(f"Total chunks created: {len(pdf_chunks)}")

    # Create FAISS index
    logger.info("Building FAISS index...")
    faiss_index, embedding_model = create_faiss_index(pdf_chunks)
    save_faiss_index(faiss_index)

    # Load the model and tokenizer
    model_path = "/root/Capstone Chatbots/mis_mod/mistral/7B-Instruct-v0.3"
    tokenizer = AutoTokenizer.from_pretrained(model_path)
    model = AutoModelForCausalLM.from_pretrained(
        model_path,
        torch_dtype=torch.float16 if torch.cuda.is_available() else torch.float32,
        device_map="auto"
    )

    # print("\n=== Chatbot is ready! ===")
    logger.info("\n=== Chatbot is ready! ===")
    qa_pairs = []
    
    try:
        while True:
            user_query = input("\nEnter your question (or type 'exit' to quit): ")
            if user_query.lower() == "exit":
                break
            
            context, answer = generate_context_aware_response(user_query, faiss_index, pdf_chunks, embedding_model, model, tokenizer)
            logger.info(f"\nContext: {context[:200]}...")  # Print first 200 chars of context
            logger.info(f"\nAnswer: {answer}")
            
            # Store Q&A pair
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            qa_pairs.append((timestamp, user_query, context, answer))
    
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
    
    finally:
        # Save Q&A pairs to Excel
        if qa_pairs:
            save_qa_to_excel(qa_pairs)
        else:
            print("No Q&A pairs to save.")

    print("Thank you for using the chatbot.")

    # Calculate and log total execution time
    end_time = time.time()
    total_time = end_time - start_time
    logger.info(f"Total execution time: {total_time:.2f} seconds")