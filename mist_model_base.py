import torch
from transformers import AutoModelForCausalLM, AutoTokenizer
import logging
import time
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import os
import re

# Initialize logging
log_filename = "chatbot_m.log"
logging.basicConfig(filename=log_filename, level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger()

start_time = time.time()

def generate_response(query, model, tokenizer, max_new_tokens=1000, temperature=0.7, top_p=0.9):
    logger.info(f"Received query: {query}")
    
    prompt = f"Human: {query}\n\nAssistant:"
    
    inputs = tokenizer(prompt, return_tensors="pt").to(model.device)
    outputs = model.generate(
        **inputs,
        max_new_tokens=max_new_tokens,
        temperature=temperature,
        top_p=top_p,
        do_sample=True,
        eos_token_id=tokenizer.eos_token_id,
        pad_token_id=tokenizer.eos_token_id
    )
    
    response = tokenizer.decode(outputs[0], skip_special_tokens=True)
    answer = response.split("Assistant:")[-1].strip()
    
    logger.info(f"Generated answer: {answer[:100]}...")  # Log first 100 chars of answer
    return answer

def sanitize_for_excel(text):
    """Remove or replace characters that Excel doesn't support."""
    # Remove control characters
    text = ''.join(char for char in text if ord(char) >= 32)
    # Replace other problematic characters
    text = re.sub(r'[\000-\010]|[\013-\014]|[\016-\031]', '', text)
    # Limit the length of the text (Excel has a 32,767 character limit per cell)
    return text[:32000]  # Leave some margin for safety

def save_qa_to_excel(qa_pairs, filename="qa_mist.xlsx"):
    if os.path.exists(filename):
        # File exists, load it and append data
        wb = load_workbook(filename)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        # File doesn't exist, create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Q&A Log"
        # Add headers
        headers = ["Timestamp", "Question", "Answer"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        start_row = 2

    # Add Q&A pairs
    for row, (timestamp, question, answer) in enumerate(qa_pairs, start=start_row):
        ws.cell(row=row, column=1, value=timestamp)
        ws.cell(row=row, column=2, value=sanitize_for_excel(question))
        # ws.cell(row=row, column=3, value=sanitize_for_excel(context))
        ws.cell(row=row, column=3, value=sanitize_for_excel(answer))

    # Save without adjusting column widths
    wb.save(filename)
    print(f"Q&A pairs saved to {filename}")

if __name__ == "__main__":
    # Load the model and tokenizer from local storage
    model_path = "/root/Capstone Chatbots/mis_mod/mistral/7B-Instruct-v0.3"
    tokenizer = AutoTokenizer.from_pretrained(model_path)
    model = AutoModelForCausalLM.from_pretrained(
        model_path,
        torch_dtype=torch.float16 if torch.cuda.is_available() else torch.float32,
        device_map="auto"
    )

    logger.info("\n=== Chatbot is ready! ===")
    qa_pairs = []

    try:
        while True:
            user_query = input("\nEnter your question (or type 'exit' to quit): ")
            if user_query.lower() == "exit":
                break

            answer = generate_response(user_query, model, tokenizer)
            print(f"\nAnswer: {answer}")

            # Store Q&A pair
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            qa_pairs.append((timestamp, user_query, answer))

    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
    finally:
        # Save Q&A pairs to Excel
        if qa_pairs:
            save_qa_to_excel(qa_pairs)
        else:
            print("No Q&A pairs to save.")

    print("Thank you for using the chatbot.")

    # Calculate and log total execution time
    end_time = time.time()
    total_time = end_time - start_time
    logger.info(f"Total execution time: {total_time:.2f} seconds")